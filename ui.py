import streamlit as st
from langchain_mistralai import ChatMistralAI
from tavily import TavilyClient
from dotenv import load_dotenv
import os
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# PAGE CONFIG


st.set_page_config(
    page_title="AI Powered Recruitment Research Assistant",
    layout="centered"
)


# LOAD ENV


load_dotenv()
MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY")
TAVILY_API_KEY  = os.getenv("TAVILY_API_KEY")


# CSS


st.markdown("""
<style>
    .main .block-container { max-width: 820px; padding-top: 2rem; }

    .app-title {
        text-align: center; font-size: 1.75rem;
        font-weight: 700; letter-spacing: -0.02em; padding-bottom: 0.3rem;
    }
    .app-sub { text-align: center; color: #888; font-size: 0.92rem; margin-bottom: 2rem; }

    /* Card wrapper */
    .result-card {
        border: 1px solid #e4e4e4; border-radius: 14px;
        overflow: hidden; margin-top: 1rem; background: #fff;
    }
    .card-header {
        background: #f7f8fa; padding: 1rem 1.4rem;
        border-bottom: 1px solid #e8e8e8;
        display: flex; align-items: center; gap: 14px;
    }
    .avatar {
        width: 52px; height: 52px; border-radius: 50%;
        background: #ddeaff; display: flex; align-items: center;
        justify-content: center; font-size: 1.15rem;
        font-weight: 700; color: #2563eb; flex-shrink: 0;
    }
    .card-name  { font-size: 1.1rem; font-weight: 700; color: #111; }
    .card-badge {
        display: inline-block; font-size: 0.7rem; padding: 2px 10px;
        border-radius: 20px; background: #eff6ff; color: #2563eb;
        font-weight: 600; margin-top: 4px;
    }

    /* Section heading inside card */
    .section-heading {
        font-size: 0.7rem; font-weight: 700; color: #2563eb;
        text-transform: uppercase; letter-spacing: 0.08em;
        padding: 10px 1.4rem 6px; background: #f0f6ff;
        border-top: 1px solid #dbeafe; border-bottom: 1px solid #dbeafe;
    }

    /* Info table */
    .info-table { width: 100%; border-collapse: collapse; }
    .info-table td {
        padding: 9px 1.4rem; border-bottom: 1px solid #f3f3f3;
        vertical-align: top; font-size: 0.84rem;
    }
    .info-table tr td:first-child {
        width: 40%; font-weight: 600; color: #374151;
        border-right: 1px solid #f3f3f3; background: #fafafa;
    }
    .info-table tr td:last-child { color: #111; word-break: break-word; }
    .info-table .na { color: #c0c0c0; font-style: italic; }
    .info-table a { color: #2563eb; text-decoration: none; }

    /* Tags */
    .tags-block { padding: 0.8rem 1.4rem; border-top: 1px solid #f0f0f0; }
    .tags-label {
        font-size: 0.68rem; color: #aaa; text-transform: uppercase;
        letter-spacing: 0.06em; margin-bottom: 8px; font-weight: 600;
    }
    .tags { display: flex; flex-wrap: wrap; gap: 6px; }
    .tag {
        font-size: 0.78rem; padding: 3px 12px; border-radius: 20px;
        background: #f0f4ff; color: #2563eb; border: 1px solid #ddeaff;
    }

    #MainMenu, footer, .stDeployButton { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# SESSION STATE


for k, v in [("mode", "Company"), ("result", None)]:
    if k not in st.session_state:
        st.session_state[k] = v


# HEADER


st.markdown('<div class="app-title"> AI Powered Recruitment Research Assistant</div>', unsafe_allow_html=True)
st.markdown('<div class="app-sub">Search detailed intelligence about companies and colleges</div>', unsafe_allow_html=True)


# API KEY GUARD


if not MISTRAL_API_KEY or not TAVILY_API_KEY:
    st.error("⚠️ API keys missing. Create a `.env` file with:")
    st.code("MISTRAL_API_KEY=your_key\nTAVILY_API_KEY=your_key", language="bash")
    st.stop()


# MODE BUTTONS

c1, c2 = st.columns(2)
with c1:
    if st.button("🏢  Company", use_container_width=True,
                 type="primary" if st.session_state.mode == "Company" else "secondary"):
        st.session_state.mode   = "Company"
        st.session_state.result = None
        st.rerun()
with c2:
    if st.button("🎓  College", use_container_width=True,
                 type="primary" if st.session_state.mode == "College" else "secondary"):
        st.session_state.mode   = "College"
        st.session_state.result = None
        st.rerun()

st.markdown("<br>", unsafe_allow_html=True)
mode = st.session_state.mode


# SEARCH FORM


with st.form("search_form"):
    entity_name = st.text_input(
        "entity", label_visibility="collapsed",
        placeholder=f"Enter {mode} name and press Search…"
    )
    submitted = st.form_submit_button("🔍  Search", use_container_width=True, type="primary")


# PROMPTS


PROMPT_COMPANY = """You are a Business Intelligence Assistant.
Entity Type: Company
Entity Name: {name}

Use ONLY the web search data provided below.
If any field is unavailable write exactly: Not Found
Return ONLY the fields below, one per line, no bullet points, no extra text.

Company Name: {name}
Company Overview:
Registered Office Address:
Industry Type:
Sub-Industry:
Headquarters and Branch Locations:
GST Registration Number:
Date of Establishment:
GST Status:
PAN Number:
CIN or UID Number:
Company Constitution:
Directors and DIN Numbers:
Employee Strength:
Revenue and Business Growth:
Official Website:
LinkedIn URL:
Current Hiring Status:
Active Job Openings:
Hiring or Expansion Plans:
HR Manager or Talent Acquisition Manager:
Contact Details:
LinkedIn Profile of HR Manager:
Recruitment Activity Level:

WEB SEARCH DATA:
{context}"""

PROMPT_COLLEGE = """You are a Business Intelligence Assistant.
Entity Type: College
Entity Name: {name}

Use ONLY the web search data provided below.
If any field is unavailable write exactly: Not Found
Return ONLY the fields below, one per line, no bullet points, no extra text.

College Name: {name}
Website:
Email:
Contact Number:
Address:
City:
Country:
University Affiliation:
Accreditation:
Established Year:
Student Count:
Faculty Count:
Courses Offered:
Departments:
College Description:

WEB SEARCH DATA:
{context}"""


# COMPANY FIELD SECTIONS (for UI grouping)


COMPANY_SECTIONS = {
    "🏢 Basic Information": [
        "Company Name",
        "Company Overview",
        "Industry Type",
        "Sub-Industry",
        "Company Constitution",
        "Date of Establishment",
        "Official Website",
    ],
    "📍 Location": [
        "Registered Office Address",
        "Headquarters and Branch Locations",
    ],
    "📋 Legal & Compliance": [
        "GST Registration Number",
        "GST Status",
        "PAN Number",
        "CIN or UID Number",
        "Directors and DIN Numbers",
    ],
    "📊 Business Metrics": [
        "Employee Strength",
        "Revenue and Business Growth",
        "LinkedIn URL",
    ],
    "💼 Hiring & HR": [
        "Current Hiring Status",
        "Active Job Openings",
        "Hiring or Expansion Plans",
        "HR Manager or Talent Acquisition Manager",
        "Contact Details",
        "LinkedIn Profile of HR Manager",
        "Recruitment Activity Level",
    ],
}

COLLEGE_SECTIONS = {
    "🎓 Basic Information": [
        "College Name",
        "University Affiliation",
        "Accreditation",
        "Established Year",
        "Website",
    ],
    "📍 Contact & Location": [
        "Email",
        "Contact Number",
        "Address",
        "City",
        "Country",
    ],
    "👥 People": [
        "Student Count",
        "Faculty Count",
    ],
}


# HELPER FUNCTIONS


def build_context(results):
    ctx = ""
    for i, r in enumerate(results.get("results", []), 1):
        ctx += (
            f"\nSource {i}\n"
            f"Title: {r.get('title','')}\n"
            f"URL: {r.get('url','')}\n"
            f"Content: {r.get('content','')}\n"
            + "-" * 60 + "\n"
        )
    return ctx


def build_prompt(mode, name, context):
    t = PROMPT_COMPANY if mode == "Company" else PROMPT_COLLEGE
    return t.format(name=name, context=context)


def parse_llm_output(text):
    """Robust parser: handles URLs, bold markdown, multi-line values."""
    data = {}
    current_key = None
    current_val = []

    for raw in text.split("\n"):
        line = raw.strip()
        line = re.sub(r"\*\*(.+?)\*\*", r"\1", line)   # strip **bold**
        line = re.sub(r"^[-•*]\s*", "", line)            # strip bullets

        m = re.match(r"^([A-Za-z][A-Za-z0-9 /\-&()]{1,60}):\s*(.*)", line)
        if m:
            if current_key is not None:
                joined = " ".join(current_val).strip()
                data[current_key] = joined if joined else "Not Found"
            current_key = m.group(1).strip()
            val = m.group(2).strip()
            current_val = [val] if val else []
        elif current_key and line:
            current_val.append(line)

    if current_key is not None:
        joined = " ".join(current_val).strip()
        data[current_key] = joined if joined else "Not Found"

    return data


def safe(val):
    if not val:
        return "Not Found"
    v = str(val).strip()
    return v if v else "Not Found"


def val_html(v):
    s = safe(v)
    if s == "Not Found":
        return '<span class="na">Not Found</span>'
    if s.lower().startswith("http"):
        return f'<a href="{s}" target="_blank">{s}</a>'
    return s


def render_card(data, mode, name):
    initials = "".join(w[0].upper() for w in name.split()[:2])
    sections = COMPANY_SECTIONS if mode == "Company" else COLLEGE_SECTIONS

    sections_html = ""
    for section_title, keys in sections.items():
        rows = ""
        for key in keys:
            val = data.get(key, "")
            rows += f"<tr><td>{key}</td><td>{val_html(val)}</td></tr>"
        sections_html += (
            f'<div class="section-heading">{section_title}</div>'
            f'<table class="info-table"><tbody>{rows}</tbody></table>'
        )

    # Tag lists for College
    tags_html = ""
    if mode == "College":
        for lbl in ["Courses Offered", "Departments"]:
            val = safe(data.get(lbl, ""))
            if val != "Not Found":
                items = [x.strip() for x in re.split(r"[,\n;]+", val) if x.strip()]
                if items:
                    tags = "".join(f'<span class="tag">{it}</span>' for it in items)
                    tags_html += (
                        f'<div class="tags-block">'
                        f'<div class="tags-label">{lbl}</div>'
                        f'<div class="tags">{tags}</div>'
                        f'</div>'
                    )

    return (
        f'<div class="result-card">'
        f'<div class="card-header">'
        f'<div class="avatar">{initials}</div>'
        f'<div><div class="card-name">{name}</div>'
        f'<span class="card-badge">{mode}</span></div>'
        f'</div>'
        f'{sections_html}'
        f'{tags_html}'
        f'</div>'
    )



# EXCEL EXPORT


def build_excel(data, mode, name):
    wb = Workbook()
    ws = wb.active
    ws.title = mode

    thin   = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Title
    ws.merge_cells("A1:B1")
    ws["A1"] = f"{mode} Intelligence Report — {name}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1E3A5F")
    ws["A1"].fill      = PatternFill("solid", fgColor="DBEAFE")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    # Column headers
    ws["A2"] = "Field"
    ws["B2"] = "Value"
    for cell in [ws["A2"], ws["B2"]]:
        cell.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2563EB")
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[2].height = 22

    sections = COMPANY_SECTIONS if mode == "Company" else COLLEGE_SECTIONS

    # Extra fields for college
    if mode == "College":
        sections = dict(sections)
        sections["📚 Academics"] = ["Courses Offered", "Departments"]
        desc_key = "College Description"
    else:
        desc_key = None

    row_idx = 3
    section_fill  = PatternFill("solid", fgColor="EFF6FF")
    label_fill    = PatternFill("solid", fgColor="F9FAFB")
    section_font  = Font(name="Arial", bold=True, size=10, color="1D4ED8")
    label_font    = Font(name="Arial", bold=True, size=10, color="374151")
    value_font    = Font(name="Arial", size=10, color="111111")

    for section_title, keys in sections.items():
        # Section header row
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        cell = ws.cell(row=row_idx, column=1, value=section_title)
        cell.font      = section_font
        cell.fill      = section_fill
        cell.alignment = left
        cell.border    = border
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

        for key in keys:
            val = safe(data.get(key, ""))
            a = ws.cell(row=row_idx, column=1, value=key)
            a.font = label_font; a.fill = label_fill
            a.alignment = left; a.border = border

            b = ws.cell(row=row_idx, column=2, value=val)
            b.font = value_font
            b.alignment = left; b.border = border

            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

    # Description at bottom
    if desc_key:
        desc = safe(data.get(desc_key, ""))
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        h = ws.cell(row=row_idx, column=1, value="Description")
        h.font = section_font; h.fill = section_fill
        h.alignment = left; h.border = border
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        d = ws.cell(row=row_idx, column=1, value=desc)
        d.font = value_font
        d.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        d.border = border
        ws.row_dimensions[row_idx].height = 60

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 58

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()



# SEARCH EXECUTION


if submitted:
    name_input = entity_name.strip()
    if not name_input:
        st.warning(f"Please enter a {mode} name.")
    else:
        with st.spinner(f"Searching for **{name_input}** …"):
            try:
                llm    = ChatMistralAI(model="mistral-large-latest", temperature=0.1, api_key=MISTRAL_API_KEY)
                client = TavilyClient(api_key=TAVILY_API_KEY)

                # Multiple targeted queries
                if mode == "Company":
                    queries = [
                        f"{name_input} official website overview",
                        f"{name_input} GSTIN GST status CIN PAN number",
                        f"{name_input} directors DIN number",
                        f"{name_input} employee count workforce revenue turnover",
                        f"{name_input} linkedin company page",
                        f"{name_input} HR manager talent acquisition hiring jobs",
                        f"{name_input} headquarters branch locations address contact",
                    ]
                else:
                    queries = [
                        f"{name_input} official website contact email address",
                        f"{name_input} courses offered departments",
                        f"{name_input} student count faculty affiliation accreditation",
                    ]

                all_results = {"results": []}
                for q in queries:
                    r = client.search(query=q, search_depth="advanced", max_results=5)
                    all_results["results"].extend(r.get("results", []))

                context  = build_context(all_results)
                prompt   = build_prompt(mode, name_input, context)
                response = llm.invoke(prompt)
                parsed   = parse_llm_output(response.content)

                st.session_state.result = {
                    "name":   name_input,
                    "mode":   mode,
                    "raw":    response.content,
                    "parsed": parsed,
                }

            except Exception as e:
                st.error(f"Error: {e}")


# DISPLAY RESULT


if st.session_state.result:
    r = st.session_state.result

    st.markdown(
        render_card(r["parsed"], r["mode"], r["name"]),
        unsafe_allow_html=True
    )

    st.markdown("<br>", unsafe_allow_html=True)

    xlsx_bytes = build_excel(r["parsed"], r["mode"], r["name"])
    st.download_button(
        label="⬇️ Download Excel Report (.xlsx)",
        data=xlsx_bytes,
        file_name=f"{r['name'].replace(' ', '_')}_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary"
    )
